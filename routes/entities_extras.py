"""V24.2 SUPABASE-ONLY: entity notes, deal documents, low-stock alerts, audit CSV, XLSX offer export."""
from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone

from flask import Blueprint, jsonify, request, session, send_file, Response

from utils import login_required, log_audit
import supabase_store as store

entities_extras_bp = Blueprint('entities_extras_bp', __name__)


VALID_ENTITY_TYPES = {'partner', 'deal', 'offer', 'product'}


def _iso_now():
    return datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')


# ==========================================================
# 1. ENTITY NOTES
# ==========================================================

@entities_extras_bp.route('/api/notes/<entity_type>/<entity_id>', methods=['GET'])
@login_required
def notes_list(entity_type, entity_id):
    if entity_type not in VALID_ENTITY_TYPES:
        return jsonify({'error': 'invalid_entity_type'}), 400
    from data_layer import select as _dl_select
    rows = _dl_select('entity_notes',
                      filters={'entity_type': entity_type, 'entity_id': entity_id},
                      order=['-pinned', '-created_at'], limit=200) or []
    return jsonify({
        'entity_type': entity_type, 'entity_id': entity_id,
        'notes': [{'id': r.get('id'), 'body': r.get('body'),
                   'created_by': r.get('created_by'), 'created_at': r.get('created_at'),
                   'pinned': bool(r.get('pinned'))} for r in rows],
    })


@entities_extras_bp.route('/api/notes/<entity_type>/<entity_id>', methods=['POST'])
@login_required
def notes_create(entity_type, entity_id):
    if entity_type not in VALID_ENTITY_TYPES:
        return jsonify({'error': 'invalid_entity_type'}), 400
    body = str((request.get_json(silent=True) or {}).get('body', '') or '').strip()
    if not body or len(body) < 2:
        return jsonify({'error': 'body_required'}), 400
    if len(body) > 5000:
        return jsonify({'error': 'body_too_long'}), 400
    nid = str(uuid.uuid4())
    from data_layer import insert as _dl_insert
    _dl_insert('entity_notes', {
        'id': nid, 'entity_type': entity_type, 'entity_id': entity_id,
        'body': body[:5000], 'created_by': session.get('username') or 'system',
        'created_at': _iso_now(), 'pinned': False,
    })
    log_audit('CREATE', entity_type, f'Note added to {entity_type}/{entity_id}')
    return jsonify({'id': nid, 'ok': True})


@entities_extras_bp.route('/api/notes/<note_id>', methods=['DELETE'])
@login_required
def notes_delete(note_id):
    from data_layer import select_one as _dl_select_one, delete as _dl_delete
    r = _dl_select_one('entity_notes', {'id': note_id})
    if not r:
        return jsonify({'error': 'note_not_found'}), 404
    is_owner = r.get('created_by') == session.get('username')
    is_admin = session.get('role') == 'admin'
    if not (is_owner or is_admin):
        return jsonify({'error': 'forbidden'}), 403
    _dl_delete('entity_notes', {'id': note_id})
    return jsonify({'deleted': 1})


@entities_extras_bp.route('/api/notes/<note_id>/pin', methods=['POST'])
@login_required
def notes_pin(note_id):
    pinned = bool((request.get_json(silent=True) or {}).get('pinned', True))
    from data_layer import update as _dl_update
    n = _dl_update('entity_notes', {'id': note_id}, {'pinned': pinned})
    return jsonify({'updated': len(n) if isinstance(n, list) else int(bool(n)),
                    'pinned': pinned})


# ==========================================================
# 2. DEAL DOCUMENTS
# ==========================================================

@entities_extras_bp.route('/api/deals/<deal_id>/docs', methods=['GET'])
@login_required
def deal_docs_list(deal_id):
    from data_layer import select as _dl_select
    rows = _dl_select('deal_documents', filters={'deal_id': deal_id},
                      order='-uploaded_at') or []
    return jsonify({
        'deal_id': deal_id,
        'documents': [{
            'id': r.get('id'), 'file_url': r.get('file_url'),
            'filename': r.get('filename'), 'doc_kind': r.get('doc_kind'),
            'size_bytes': r.get('size_bytes'),
            'uploaded_by': r.get('uploaded_by'), 'uploaded_at': r.get('uploaded_at'),
            'note': r.get('note'),
        } for r in rows]
    })


@entities_extras_bp.route('/api/deals/<deal_id>/docs', methods=['POST'])
@login_required
def deal_docs_attach(deal_id):
    body = request.get_json(silent=True) or {}
    file_url = str(body.get('file_url') or '').strip()
    filename = str(body.get('filename') or '').strip()
    if not file_url or not filename:
        return jsonify({'error': 'file_url_and_filename_required'}), 400
    doc_kind = str(body.get('doc_kind') or 'other').strip().lower()[:40]
    size_bytes = int(body.get('size_bytes') or 0)
    note = str(body.get('note') or '').strip()[:500]
    did = str(uuid.uuid4())
    # Verifikuj da deal postoji
    if not store.get_entity('deals', deal_id):
        return jsonify({'error': 'deal_not_found'}), 404
    from data_layer import insert as _dl_insert
    _dl_insert('deal_documents', {
        'id': did, 'deal_id': deal_id, 'file_url': file_url,
        'filename': filename[:200], 'doc_kind': doc_kind,
        'size_bytes': size_bytes, 'uploaded_by': session.get('username') or 'system',
        'uploaded_at': _iso_now(), 'note': note,
    })
    log_audit('CREATE', 'deal_documents',
              f'Attached "{filename}" ({doc_kind}) to deal {deal_id}')
    return jsonify({'id': did, 'ok': True})


@entities_extras_bp.route('/api/deals/docs/<doc_id>', methods=['DELETE'])
@login_required
def deal_docs_detach(doc_id):
    from data_layer import delete as _dl_delete
    n = _dl_delete('deal_documents', {'id': doc_id})
    return jsonify({'deleted': int(n or 0)})


# ==========================================================
# 3. LOW-STOCK ALERTS
# ==========================================================

@entities_extras_bp.route('/api/inventory/low-stock', methods=['GET'])
@login_required
def inventory_low_stock():
    """Vraca sve partner_inventory redove gde je qty_free <= threshold."""
    try:
        threshold = float(request.args.get('threshold', 0))
    except ValueError:
        threshold = 0
    from data_layer import select as _dl_select
    all_rows = _dl_select('partner_inventory', limit=5000) or []
    # PostgREST nema izraz "col1 - col2 <= X" u REST filteru; filtriraj u Pythonu
    filtered = []
    for r in all_rows:
        free = float(r.get('qty_on_hand') or 0) - float(r.get('qty_reserved') or 0)
        if free <= threshold:
            filtered.append((free, r))
    filtered.sort(key=lambda x: (x[0], x[1].get('last_movement_at') or ''))
    filtered = filtered[:500]

    # Enrich sa partner + product names
    partner_cache = {p.get('id'): p for p in store.list_entities('partners')}
    product_cache = {p.get('id'): p for p in store.list_entities('products')}
    items = []
    for free, r in filtered:
        pid = r.get('partner_id'); pdid = r.get('product_id')
        partner_name = (partner_cache.get(pid) or {}).get('companyName')
        product_name = (product_cache.get(pdid) or {}).get('name')
        items.append({
            'id': r.get('id'), 'partner_id': pid, 'partner_name': partner_name,
            'product_id': pdid, 'product_name': product_name,
            'qty_on_hand': float(r.get('qty_on_hand') or 0),
            'qty_reserved': float(r.get('qty_reserved') or 0),
            'qty_free': free, 'unit': r.get('unit'),
            'last_movement_at': r.get('last_movement_at'),
            'severity': 'critical' if free <= 0 else ('warning' if free <= threshold/2 else 'notice'),
        })
    return jsonify({'items': items, 'total': len(items), 'threshold': threshold})


# ==========================================================
# 4. AUDIT LOG CSV EXPORT
# ==========================================================

@entities_extras_bp.route('/api/audit/export.csv', methods=['GET'])
@login_required
def audit_export_csv():
    """V24.2 SUPABASE-ONLY: audit CSV export."""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only.'}), 403
    date_from = (request.args.get('from') or '').strip()
    date_to = (request.args.get('to') or '').strip()
    filters = {}
    if date_from: filters['timestamp'] = ('gte', date_from)
    # NAPOMENA: PostgREST prima jedan op po koloni preko tuple. Za AND
    # (>= from AND <= to), koristimo prvi kroz filter a drugi filtriramo lokalno.
    from data_layer import select as _dl_select
    rows = _dl_select('audit_logs', filters=filters, order='-timestamp',
                      limit=100000) or []
    if date_to:
        rows = [r for r in rows if str(r.get('timestamp') or '') <= date_to]

    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(['timestamp', 'action', 'module', 'username', 'details', 'ip_address', 'is_suspicious'])
    for r in rows:
        w.writerow([
            r.get('timestamp'), r.get('action'), r.get('module'),
            r.get('username'), (r.get('details') or '')[:500],
            r.get('ip_address'), 1 if r.get('is_suspicious') else 0
        ])
    log_audit('READ', 'audit', f'CSV export by {session.get("username")}: {len(rows)} rows',
              is_suspicious=False)
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="audit-log-{_iso_now()[:10]}.csv"'}
    )


# ==========================================================
# 5. XLSX OFFER EXPORT
# ==========================================================

@entities_extras_bp.route('/api/offers/<offer_id>/export.xlsx', methods=['GET'])
@login_required
def offer_export_xlsx(offer_id):
    offer = store.get_entity('offers', offer_id)
    if not offer:
        return jsonify({'error': 'offer_not_found'}), 404

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback: CSV export
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['Offer', offer.get('offerNo', offer_id)])
        w.writerow([])
        w.writerow(['#', 'Product', 'Qty', 'Unit', 'Unit Price', 'Total'])
        for i, item in enumerate(offer.get('items', []) or [], 1):
            w.writerow([
                i, item.get('name', ''),
                item.get('quantity', 0), item.get('unit', ''),
                item.get('unitPrice', 0),
                (item.get('quantity', 0) or 0) * (item.get('unitPrice', 0) or 0),
            ])
        return Response(
            buf.getvalue(), mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="offer-{offer.get("offerNo", offer_id)}.csv"'}
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Offer {str(offer.get("offerNo", offer_id))[:20]}'
    ws['A1'] = 'Aspidus — Offer'; ws['A1'].font = Font(size=16, bold=True)
    ws['A3'] = 'Offer No:'; ws['B3'] = offer.get('offerNo', '')
    ws['A4'] = 'Date:';     ws['B4'] = offer.get('date', offer.get('createdAt', ''))
    ws['A5'] = 'Customer:'; ws['B5'] = offer.get('customerName', '')
    ws['A6'] = 'Currency:'; ws['B6'] = offer.get('currency', 'USD')

    headers = ['#', 'Product', 'Description', 'Qty', 'Unit', 'Unit Price', 'Total']
    row = 8
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    grand_total = 0
    for i, item in enumerate(offer.get('items', []) or [], 1):
        row += 1
        qty = float(item.get('quantity', 0) or 0)
        price = float(item.get('unitPrice', 0) or 0)
        total = qty * price
        grand_total += total
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get('name', ''))
        ws.cell(row=row, column=3, value=item.get('description', ''))
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=item.get('unit', ''))
        ws.cell(row=row, column=6, value=price)
        ws.cell(row=row, column=7, value=total)

    row += 2
    ws.cell(row=row, column=6, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=7, value=grand_total).font = Font(bold=True, size=14)
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'offer-{offer.get("offerNo", offer_id)}.xlsx',
    )
